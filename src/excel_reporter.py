#!/usr/bin/env python3
"""
excel_reporter.py - Gerador de Relatórios Excel EXPANDIDO
Sistema de Análise Estatística DL 54/2018 v2.0

VERSÃO EXPANDIDA: 15+ sheets com análises detalhadas

Autor: Sistema de Qualidade de Dados Educacionais
"""

import pandas as pd
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from config import ConfigManager
from utils import format_percentage, format_number, create_timestamp


logger = logging.getLogger('DL54.excel_reporter')


class ExcelReporter:
    """
    Gerador de relatórios Excel profissionais EXPANDIDO.
    
    Características:
    - 15+ sheets temáticos
    - Formatação profissional
    - Fórmulas automáticas
    - Cores por categoria
    - Validação de dados
    """
    
    def __init__(self, config: ConfigManager, logger_instance: logging.Logger):
        """Inicializa gerador de relatórios Excel expandido."""
        self.config = config
        self.logger = logger_instance
        
        self.report_config = config.get_report_config()
        self.escola_mapping = config.get_escola_mapping()
        self.colors = config.colors
        
        self.output_dir = Path(config.get('IO', 'OUTPUT_DIR'))
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Estilos
        self.header_fill = PatternFill(start_color=self.colors.primary.replace('#', ''),
                                       end_color=self.colors.primary.replace('#', ''),
                                       fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_report(self, stats: Dict[str, Any],
                       raw_data: Optional[pd.DataFrame] = None) -> Path:
        """Gera relatório Excel completo EXPANDIDO."""
        self.logger.info("=" * 80)
        self.logger.info("GERANDO RELATÓRIO EXCEL EXPANDIDO")
        self.logger.info("=" * 80)
        
        timestamp = create_timestamp()
        filename = self.report_config['xlsx_filename'].replace('.xlsx', f'_EXPANDIDO_{timestamp}.xlsx')
        output_path = self.output_dir / filename
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # ========================================
            # SHEET 1: Resumo Executivo
            # ========================================
            self._create_executive_summary(writer, stats)
            
            # ========================================
            # SHEET 2: Análise Global
            # ========================================
            if 'global' in stats:
                self._create_global_sheet(writer, stats['global'])
            
            # ========================================
            # SHEET 3: Por Escola
            # ========================================
            if 'por_escola' in stats:
                self._create_school_sheet(writer, stats['por_escola'])
            
            # ========================================
            # SHEET 4: Por Ano
            # ========================================
            if 'por_ano' in stats:
                self._create_year_sheet(writer, stats['por_ano'])
            
            # ========================================
            # SHEET 5: Por Turma
            # ========================================
            if 'por_turma' in stats:
                self._create_turma_sheet(writer, stats['por_turma'])
            
            # ========================================
            # SHEET 6: Por Ano e Turma
            # ========================================
            if 'por_ano_turma' in stats:
                self._create_ano_turma_sheet(writer, stats['por_ano_turma'])
            
            # ========================================
            # SHEET 7: Estatísticas por Aluno
            # ========================================
            if 'estatisticas_aluno_turma' in stats:
                self._create_estatisticas_aluno_sheet(writer, stats['estatisticas_aluno_turma'])
            
            # ========================================
            # SHEET 8: Alíneas Detalhadas
            # ========================================
            if 'alineas_detalhadas' in stats:
                self._create_alineas_sheet(writer, stats['alineas_detalhadas'])
            
            # ========================================
            # SHEET 9: Terapias
            # ========================================
            if 'terapias_completas' in stats:
                self._create_terapias_sheet(writer, stats['terapias_completas'])
            
            # ========================================
            # SHEET 10: Por Sexo
            # ========================================
            if 'por_sexo' in stats:
                self._create_gender_sheet(writer, stats['por_sexo'])
            
            # ========================================
            # SHEET 11: Sexo Detalhado
            # ========================================
            if 'sexo_detalhado' in stats:
                self._create_sexo_detalhado_sheet(writer, stats['sexo_detalhado'])
            
            # ========================================
            # SHEET 12: Por Escalão ASE
            # ========================================
            if 'por_escalao_ase' in stats:
                self._create_ase_sheet(writer, stats['por_escalao_ase'])
            
            # ========================================
            # SHEET 13: Rankings
            # ========================================
            if 'rankings' in stats:
                self._create_rankings_sheet(writer, stats['rankings'])
            
            # ========================================
            # SHEET 14: Comparações
            # ========================================
            if 'comparacoes' in stats:
                self._create_comparisons_sheet(writer, stats['comparacoes'])
            
            # ========================================
            # SHEET 15: Dados Brutos
            # ========================================
            if raw_data is not None:
                self._create_raw_data_sheet(writer, raw_data)
        
        # Aplicar formatação
        self._apply_formatting(output_path)
        
        self.logger.info(f"✓ Relatório Excel EXPANDIDO gerado: {output_path.name}")
        self.logger.info("=" * 80)
        
        return output_path
    
    def _create_executive_summary(self, writer: pd.ExcelWriter, stats: Dict):
        """SHEET 1: Resumo Executivo."""
        self.logger.info("Criando sheet: Resumo Executivo...")
        
        data = []
        
        # Cabeçalho
        data.append(['RESUMO EXECUTIVO - DL 54/2018', '', ''])
        data.append(['Data do Relatório:', datetime.now().strftime('%d/%m/%Y %H:%M'), ''])
        data.append(['', '', ''])
        
        # Métricas Globais
        if 'metadata' in stats:
            data.append(['MÉTRICAS GLOBAIS', '', ''])
            data.append(['Total de Alunos:', stats['metadata']['total_alunos'], ''])
            data.append(['Total de Colunas:', stats['metadata']['colunas'], ''])
            data.append(['Versão do Sistema:', stats['metadata'].get('versao', 'N/A'), ''])
            data.append(['', '', ''])
        
        # Medidas Principais
        if 'global' in stats:
            data.append(['DISTRIBUIÇÃO DE MEDIDAS', '', ''])
            data.append(['Medida', 'N', '%'])
            
            for medida, valores in stats['global']['medidas_principais'].items():
                data.append([medida, valores['n'], f"{valores['percentagem']:.1f}%"])
            
            data.append(['', '', ''])
        
        # Escolas
        if 'por_escola' in stats:
            data.append(['DISTRIBUIÇÃO POR ESCOLA', '', ''])
            data.append(['Escola', 'Total Alunos', '% do Agrupamento'])
            
            for codigo, escola in stats['por_escola'].items():
                nome_curto = escola['nome'].split(',')[0]
                data.append([nome_curto, escola['total_alunos'],
                            f"{escola['peso_no_agrupamento']:.1f}%"])
            
            data.append(['', '', ''])
        
        # Anos
        if 'por_ano' in stats:
            data.append(['DISTRIBUIÇÃO POR ANO', '', ''])
            data.append(['Ano', 'Total Alunos', '% do Total'])
            
            for ano in sorted(stats['por_ano'].keys()):
                ano_dados = stats['por_ano'][ano]
                data.append([ano_dados['ano_nome'], ano_dados['total_alunos'],
                            f"{ano_dados['peso_no_total']:.1f}%"])
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Resumo Executivo', index=False, header=False)
    
    def _create_global_sheet(self, writer: pd.ExcelWriter, global_stats: Dict):
        """SHEET 2: Análise Global."""
        self.logger.info("Criando sheet: Análise Global...")
        
        data = []
        
        # Medidas Principais
        data.append(['MEDIDAS PRINCIPAIS', '', '', ''])
        data.append(['Medida', 'N', '%', 'Sem Medida'])
        
        for medida, valores in global_stats['medidas_principais'].items():
            data.append([
                medida,
                valores['n'],
                f"{valores['percentagem']:.1f}%",
                valores.get('sem_medida', 0)
            ])
        
        data.append(['', '', '', ''])
        
        # Medidas Detalhadas - Universais
        data.append(['MEDIDAS UNIVERSAIS (Art. 8º)', '', '', ''])
        data.append(['Alínea', 'N', '%', ''])
        
        for alinea, valores in global_stats['medidas_detalhadas']['universais'].items():
            data.append([alinea, valores['n'], f"{valores['percentagem']:.1f}%", ''])
        
        data.append(['', '', '', ''])
        
        # Medidas Detalhadas - Seletivas
        data.append(['MEDIDAS SELETIVAS (Art. 9º)', '', '', ''])
        data.append(['Alínea', 'N', '%', ''])
        
        for alinea, valores in global_stats['medidas_detalhadas']['seletivas'].items():
            data.append([alinea, valores['n'], f"{valores['percentagem']:.1f}%", ''])
        
        data.append(['', '', '', ''])
        
        # Medidas Detalhadas - Adicionais
        data.append(['MEDIDAS ADICIONAIS (Art. 10º)', '', '', ''])
        data.append(['Alínea', 'N', '%', ''])
        
        for alinea, valores in global_stats['medidas_detalhadas']['adicionais'].items():
            data.append([alinea, valores['n'], f"{valores['percentagem']:.1f}%", ''])
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Análise Global', index=False, header=False)
    
    def _create_school_sheet(self, writer: pd.ExcelWriter, school_stats: Dict):
        """SHEET 3: Por Escola."""
        self.logger.info("Criando sheet: Por Escola...")
        
        rows = []
        
        for codigo, escola in school_stats.items():
            rows.append({
                'Código': codigo,
                'Escola': escola['nome'],
                'Total Alunos': escola['total_alunos'],
                '% do Agrupamento': f"{escola['peso_no_agrupamento']:.1f}%",
                'M. Universais (N)': escola['medidas_principais'].get('Medidas Universais', {}).get('n', 0),
                'M. Universais (%)': f"{escola['medidas_principais'].get('Medidas Universais', {}).get('percentagem', 0):.1f}%",
                'M. Seletivas (N)': escola['medidas_principais'].get('Medidas Seletivas', {}).get('n', 0),
                'M. Seletivas (%)': f"{escola['medidas_principais'].get('Medidas Seletivas', {}).get('percentagem', 0):.1f}%",
                'M. Adicionais (N)': escola['medidas_principais'].get('Medidas Adicionais', {}).get('n', 0),
                'M. Adicionais (%)': f"{escola['medidas_principais'].get('Medidas Adicionais', {}).get('percentagem', 0):.1f}%"
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Por Escola', index=False)
    
    def _create_year_sheet(self, writer: pd.ExcelWriter, year_stats: Dict):
        """SHEET 4: Por Ano."""
        self.logger.info("Criando sheet: Por Ano...")
        
        rows = []
        
        for ano in sorted(year_stats.keys()):
            ano_dados = year_stats[ano]
            rows.append({
                'Ano': ano_dados['ano_nome'],
                'Total Alunos': ano_dados['total_alunos'],
                '% do Total': f"{ano_dados['peso_no_total']:.1f}%",
                'M. Universais (N)': ano_dados['medidas_principais'].get('Medidas Universais', {}).get('n', 0),
                'M. Universais (%)': f"{ano_dados['medidas_principais'].get('Medidas Universais', {}).get('percentagem', 0):.1f}%",
                'M. Seletivas (N)': ano_dados['medidas_principais'].get('Medidas Seletivas', {}).get('n', 0),
                'M. Seletivas (%)': f"{ano_dados['medidas_principais'].get('Medidas Seletivas', {}).get('percentagem', 0):.1f}%",
                'M. Adicionais (N)': ano_dados['medidas_principais'].get('Medidas Adicionais', {}).get('n', 0),
                'M. Adicionais (%)': f"{ano_dados['medidas_principais'].get('Medidas Adicionais', {}).get('percentagem', 0):.1f}%"
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Por Ano', index=False)
    
    def _create_turma_sheet(self, writer: pd.ExcelWriter, turma_stats: Dict):
        """SHEET 5: Por Turma."""
        self.logger.info("Criando sheet: Por Turma...")
        
        rows = []
        
        for turma, dados in sorted(turma_stats.items()):
            rows.append({
                'Turma': turma,
                'Total Alunos': dados['total_alunos'],
                'M. Universais (N)': dados['medidas_principais'].get('Medidas Universais', {}).get('n', 0),
                'M. Universais (%)': f"{dados['medidas_principais'].get('Medidas Universais', {}).get('percentagem', 0):.1f}%",
                'M. Seletivas (N)': dados['medidas_principais'].get('Medidas Seletivas', {}).get('n', 0),
                'M. Seletivas (%)': f"{dados['medidas_principais'].get('Medidas Seletivas', {}).get('percentagem', 0):.1f}%",
                'M. Adicionais (N)': dados['medidas_principais'].get('Medidas Adicionais', {}).get('n', 0),
                'M. Adicionais (%)': f"{dados['medidas_principais'].get('Medidas Adicionais', {}).get('percentagem', 0):.1f}%"
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Por Turma', index=False)
    
    def _create_ano_turma_sheet(self, writer: pd.ExcelWriter, ano_turma_stats: Dict):
        """SHEET 6: Por Ano e Turma."""
        self.logger.info("Criando sheet: Por Ano e Turma...")
        
        rows = []
        
        for chave, dados in sorted(ano_turma_stats.items()):
            rows.append({
                'Ano': dados['ano'],
                'Turma': dados['turma'],
                'Descrição': dados['nome'],
                'Total Alunos': dados['total_alunos'],
                'M. Universais (N)': dados['medidas_principais'].get('Medidas Universais', {}).get('n', 0),
                'M. Universais (%)': f"{dados['medidas_principais'].get('Medidas Universais', {}).get('percentagem', 0):.1f}%",
                'M. Seletivas (N)': dados['medidas_principais'].get('Medidas Seletivas', {}).get('n', 0),
                'M. Seletivas (%)': f"{dados['medidas_principais'].get('Medidas Seletivas', {}).get('percentagem', 0):.1f}%",
                'M. Adicionais (N)': dados['medidas_principais'].get('Medidas Adicionais', {}).get('n', 0),
                'M. Adicionais (%)': f"{dados['medidas_principais'].get('Medidas Adicionais', {}).get('percentagem', 0):.1f}%"
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Por Ano e Turma', index=False)
    
    def _create_estatisticas_aluno_sheet(self, writer: pd.ExcelWriter, stats_aluno: Dict):
        """SHEET 7: Estatísticas por Aluno."""
        self.logger.info("Criando sheet: Estatísticas por Aluno...")
        
        rows = []
        
        for chave, dados in sorted(stats_aluno.items()):
            stats = dados['estatisticas_medidas_por_aluno']
            rows.append({
                'Ano': dados['ano'],
                'Turma': dados['turma'],
                'Descrição': dados['nome'],
                'N Alunos': dados['total_alunos'],
                'Máximo': stats['maximo'],
                'Mínimo': stats['minimo'],
                'Média': f"{stats['media']:.2f}",
                'Mediana': f"{stats['mediana']:.1f}",
                'Desvio Padrão': f"{stats['desvio_padrao']:.2f}",
                'P25': f"{stats['p25']:.1f}",
                'P75': f"{stats['p75']:.1f}"
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Estatísticas por Aluno', index=False)
    
    def _create_alineas_sheet(self, writer: pd.ExcelWriter, alineas_stats: Dict):
        """SHEET 8: Alíneas Detalhadas."""
        self.logger.info("Criando sheet: Alíneas Detalhadas...")
        
        rows = []
        
        for ano_nome, dados_ano in sorted(alineas_stats.items()):
            # Universais
            for alinea, valores in dados_ano['universais'].items():
                rows.append({
                    'Ano': ano_nome,
                    'Tipo': 'Universal',
                    'Alínea': alinea,
                    'N': valores['n'],
                    '%': f"{valores['percentagem']:.1f}%"
                })
            
            # Seletivas
            for alinea, valores in dados_ano['seletivas'].items():
                rows.append({
                    'Ano': ano_nome,
                    'Tipo': 'Seletiva',
                    'Alínea': alinea,
                    'N': valores['n'],
                    '%': f"{valores['percentagem']:.1f}%"
                })
            
            # Adicionais
            for alinea, valores in dados_ano['adicionais'].items():
                rows.append({
                    'Ano': ano_nome,
                    'Tipo': 'Adicional',
                    'Alínea': alinea,
                    'N': valores['n'],
                    '%': f"{valores['percentagem']:.1f}%"
                })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Alíneas Detalhadas', index=False)
    
    def _create_terapias_sheet(self, writer: pd.ExcelWriter, terapias_stats: Dict):
        """SHEET 9: Terapias."""
        self.logger.info("Criando sheet: Terapias...")
        
        data = []
        
        # Global
        data.append(['TERAPIAS - VISÃO GLOBAL', '', ''])
        data.append(['Terapia', 'N', '%'])
        
        for terapia, valores in terapias_stats['global'].items():
            data.append([terapia, valores['n'], f"{valores['percentagem']:.1f}%"])
        
        data.append(['', '', ''])
        
        # Por Ano
        data.append(['TERAPIAS POR ANO', '', ''])
        
        for ano_nome, terapias_ano in sorted(terapias_stats['por_ano'].items()):
            data.append([f'--- {ano_nome} ---', '', ''])
            for terapia, valores in terapias_ano.items():
                data.append([terapia, valores['n'], f"{valores['percentagem']:.1f}%"])
            data.append(['', '', ''])
        
        # Por Sexo
        data.append(['TERAPIAS POR SEXO', '', ''])
        
        for sexo, terapias_sexo in sorted(terapias_stats['por_sexo'].items()):
            data.append([f'--- Sexo {sexo} ---', '', ''])
            for terapia, valores in terapias_sexo.items():
                data.append([terapia, valores['n'], f"{valores['percentagem']:.1f}%"])
            data.append(['', '', ''])
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Terapias', index=False, header=False)
    
    def _create_gender_sheet(self, writer: pd.ExcelWriter, gender_stats: Dict):
        """SHEET 10: Por Sexo."""
        self.logger.info("Criando sheet: Por Sexo...")
        
        rows = []
        
        for sexo, dados in gender_stats.items():
            if sexo.startswith('_'):
                continue
            
            rows.append({
                'Sexo': sexo,
                'Total Alunos': dados['total_alunos'],
                '% do Total': f"{dados['percentagem_do_total']:.1f}%",
                'M. Universais (N)': dados['medidas_principais'].get('Medidas Universais', {}).get('n', 0),
                'M. Universais (%)': f"{dados['medidas_principais'].get('Medidas Universais', {}).get('percentagem_no_grupo', 0):.1f}%",
                'M. Seletivas (N)': dados['medidas_principais'].get('Medidas Seletivas', {}).get('n', 0),
                'M. Seletivas (%)': f"{dados['medidas_principais'].get('Medidas Seletivas', {}).get('percentagem_no_grupo', 0):.1f}%",
                'M. Adicionais (N)': dados['medidas_principais'].get('Medidas Adicionais', {}).get('n', 0),
                'M. Adicionais (%)': f"{dados['medidas_principais'].get('Medidas Adicionais', {}).get('percentagem_no_grupo', 0):.1f}%"
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Por Sexo', index=False)
    
    def _create_sexo_detalhado_sheet(self, writer: pd.ExcelWriter, sexo_stats: Dict):
        """SHEET 11: Sexo Detalhado."""
        self.logger.info("Criando sheet: Sexo Detalhado...")
        
        rows = []
        
        for sexo, dados in sorted(sexo_stats.items()):
            # Universais
            for alinea, valores in dados['medidas_detalhadas']['universais'].items():
                rows.append({
                    'Sexo': sexo,
                    'Tipo': 'Universal',
                    'Alínea': alinea,
                    'N': valores['n'],
                    '%': f"{valores['percentagem']:.1f}%"
                })
            
            # Seletivas
            for alinea, valores in dados['medidas_detalhadas']['seletivas'].items():
                rows.append({
                    'Sexo': sexo,
                    'Tipo': 'Seletiva',
                    'Alínea': alinea,
                    'N': valores['n'],
                    '%': f"{valores['percentagem']:.1f}%"
                })
            
            # Adicionais
            for alinea, valores in dados['medidas_detalhadas']['adicionais'].items():
                rows.append({
                    'Sexo': sexo,
                    'Tipo': 'Adicional',
                    'Alínea': alinea,
                    'N': valores['n'],
                    '%': f"{valores['percentagem']:.1f}%"
                })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Sexo Detalhado', index=False)
    
    def _create_ase_sheet(self, writer: pd.ExcelWriter, ase_stats: Dict):
        """SHEET 12: Por Escalão ASE."""
        self.logger.info("Criando sheet: Por Escalão ASE...")
        
        rows = []
        
        for escalao, dados in ase_stats.items():
            rows.append({
                'Escalão': escalao,
                'Total Alunos': dados['total_alunos'],
                '% do Total': f"{dados['percentagem_do_total']:.1f}%",
                'M. Universais (N)': dados['medidas_principais'].get('Medidas Universais', {}).get('n', 0),
                'M. Universais (%)': f"{dados['medidas_principais'].get('Medidas Universais', {}).get('percentagem', 0):.1f}%",
                'M. Seletivas (N)': dados['medidas_principais'].get('Medidas Seletivas', {}).get('n', 0),
                'M. Seletivas (%)': f"{dados['medidas_principais'].get('Medidas Seletivas', {}).get('percentagem', 0):.1f}%",
                'M. Adicionais (N)': dados['medidas_principais'].get('Medidas Adicionais', {}).get('n', 0),
                'M. Adicionais (%)': f"{dados['medidas_principais'].get('Medidas Adicionais', {}).get('percentagem', 0):.1f}%"
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Por Escalão ASE', index=False)
    
    def _create_rankings_sheet(self, writer: pd.ExcelWriter, rankings: Dict):
        """SHEET 13: Rankings."""
        self.logger.info("Criando sheet: Rankings...")
        
        data = []
        
        for medida, ranking_data in rankings.items():
            data.append([f'RANKING - {medida}', '', '', ''])
            data.append(['Posição', 'Escola', 'N', '%'])
            
            for item in ranking_data:
                nome_curto = item['nome'].split(',')[0]
                data.append([
                    item['rank'],
                    nome_curto,
                    item['alunos_com_medida'],
                    f"{item['percentagem']:.1f}%"
                ])
            
            data.append(['', '', '', ''])
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Rankings', index=False, header=False)
    
    def _create_comparisons_sheet(self, writer: pd.ExcelWriter, comparisons: Dict):
        """SHEET 14: Comparações."""
        self.logger.info("Criando sheet: Comparações...")
        
        rows = []
        
        for escola_nome, comp_data in comparisons.items():
            for medida, valores in comp_data.items():
                rows.append({
                    'Escola': escola_nome.split(',')[0],
                    'Medida': medida,
                    '% Escola': f"{valores['escola']:.1f}%",
                    '% Agrupamento': f"{valores['agrupamento']:.1f}%",
                    'Diferença': f"{valores['diferenca']:.1f}%",
                    'Superior': 'Sim' if valores['superior'] else 'Não'
                })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Comparações', index=False)
    
    def _create_raw_data_sheet(self, writer: pd.ExcelWriter, raw_data: pd.DataFrame):
        """SHEET 15: Dados Brutos."""
        self.logger.info("Criando sheet: Dados Brutos...")
        
        raw_data.to_excel(writer, sheet_name='Dados Brutos', index=False)
    
    def _apply_formatting(self, filepath: Path):
        """Aplica formatação profissional ao Excel."""
        self.logger.info("Aplicando formatação ao Excel...")
        
        try:
            wb = openpyxl.load_workbook(filepath)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Formatação de cabeçalhos (primeira linha)
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = self.border
                
                # Auto-ajustar largura das colunas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Congelar primeira linha
                ws.freeze_panes = 'A2'
            
            wb.save(filepath)
            self.logger.info("✓ Formatação aplicada")
            
        except Exception as e:
            self.logger.warning(f"Erro ao aplicar formatação: {e}")
